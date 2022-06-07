import cv2 as cv
import openpyxl
import time

from Settings import *
from Hash import GetImageHash

wb = openpyxl.load_workbook(HASH_EXCEL_PATH)
sheet = wb.active

def HanmingDistance(tempImgHash: str, imgHash: str) -> int:
    w = 0
    for i in range(len(tempImgHash)):
        w += bin(int(tempImgHash[i], HASH_BASE) ^ int(imgHash[i], HASH_BASE)).count("1")
    return w

def ExcRead(row: int):
    cardName = sheet.cell(row, 1).value
    dct = sheet.cell(row, 2).value
    return cardName, dct

tmpImg = {
    'hash': GetImageHash(TEMPLATE_PATH),
    'cvData': cv.imread(TEMPLATE_PATH)
}
resImg = {
    'file': "",
    'dist': -1
}

start = time.time()
# 指定要列出所有檔案的目錄
for i in range(2, sheet.max_row + 1):
    name, dct = ExcRead(i)
    dis = HanmingDistance(tmpImg["hash"], dct)

    if resImg["dist"] == -1 or dis < resImg["dist"]:
        resImg["dist"] = dis
        resImg["file"] = name
end = time.time()

print(format(end-start))
print("minimum hanming: ", resImg["dist"])
cv.imshow("Template Image", tmpImg["cvData"])
cv.imshow("Detected Image", cv.imread(IMAGE_DB_PATH + resImg["file"]))
cv.waitKey(0)
cv.destroyAllWindows() # AW bigcase